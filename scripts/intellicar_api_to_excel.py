"""
Intellicar Standard API -> single Excel workbook of every endpoint's response shape.

One sheet per endpoint (response.data flattened to a real table) plus an Index
sheet listing each call's status / row count / params. History endpoints use a
last-24h window. Per-vehicle endpoints use the first vehicle from listvehicles.

Read-only: skips immobilize / mobilize / lock / unlock / changepassword and
mosfetcontrol enable+disable.

Run:
    python scripts/intellicar_api_to_excel.py
Output:
    sample-data/intellicar_api_responses.xlsx
"""
from __future__ import annotations
import json
import os
import sys
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = "https://apiplatform.intellicar.in/api/standard"
MOSFET_BASE = "https://apiplatform.intellicar.in/api/mosfetcontrol"
OUT_DIR = Path(__file__).resolve().parent.parent / "sample-data"
OUT_FILE = OUT_DIR / "intellicar_api_responses.xlsx"
RATE_DELAY_S = 0.4  # docs: >=400 ms between calls, max 3/sec


def load_env() -> tuple[str, str]:
    env_path = Path(__file__).resolve().parent / ".env"
    if env_path.exists():
        for line in env_path.read_text().splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, _, v = line.partition("=")
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))
    user = os.environ.get("INTELLICAR_USERNAME")
    pw = os.environ.get("INTELLICAR_PASSWORD")
    if not user or not pw:
        sys.exit("ERROR: set INTELLICAR_USERNAME and INTELLICAR_PASSWORD in scripts/.env")
    return user, pw


def post(url: str, payload: dict, timeout: int = 30) -> dict:
    try:
        r = requests.post(url, json=payload, timeout=timeout)
    except requests.RequestException as exc:
        return {"status": "FAILURE", "err": "network", "msg": str(exc), "data": None}
    try:
        return r.json()
    except ValueError:
        return {"status": "FAILURE", "err": f"http_{r.status_code}",
                "msg": "non-JSON body", "data": None, "_raw_text": r.text[:500]}


def ms(dt: datetime) -> int:
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return int(dt.timestamp() * 1000)


def _flatten_one(d):
    """Flatten one level of nesting: {a:{b:1}} -> {'a.b': 1}; lists -> JSON string."""
    if not isinstance(d, dict):
        return {"value": d}
    out: dict = {}
    for k, v in d.items():
        if isinstance(v, dict):
            for k2, v2 in v.items():
                if isinstance(v2, (dict, list)):
                    out[f"{k}.{k2}"] = json.dumps(v2, ensure_ascii=False)
                else:
                    out[f"{k}.{k2}"] = v2
        elif isinstance(v, list):
            out[k] = json.dumps(v, ensure_ascii=False)
        else:
            out[k] = v
    return out


def flatten_to_rows(body) -> tuple[list[dict], str]:
    """Return (rows, note) where rows is a list of dicts ready to write as a table."""
    if not isinstance(body, dict):
        return [{"value": str(body)}], "non-dict response"
    data = body.get("data")
    if data is None:
        return [{"status": body.get("status"),
                 "err": body.get("err"),
                 "msg": body.get("msg")}], "no data field"
    if isinstance(data, list):
        if not data:
            return [], "data is empty list"
        return [_flatten_one(item) for item in data], None
    if isinstance(data, dict):
        rows = []
        for k, v in data.items():
            if isinstance(v, (dict, list)):
                rows.append({"key": k, "value": json.dumps(v, ensure_ascii=False)})
            else:
                rows.append({"key": k, "value": v})
        return rows, None
    return [{"value": str(data)}], None


def write_sheet(wb: Workbook, name: str, rows: list[dict], header_meta: list[tuple[str, str]]):
    """Create a sheet `name` with metadata header and flattened table rows."""
    # Excel sheet names: max 31 chars, no : \ / ? * [ ]
    safe = name.replace("/", "_")[:31]
    ws = wb.create_sheet(safe)
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="DDEBF7")

    # metadata block
    for i, (k, v) in enumerate(header_meta, start=1):
        ws.cell(row=i, column=1, value=k).font = bold
        ws.cell(row=i, column=2, value=v)
    start_row = len(header_meta) + 2

    if not rows:
        ws.cell(row=start_row, column=1, value="(no rows)").font = bold
        ws.column_dimensions["A"].width = 30
        return

    # union of keys preserving first-seen order
    cols: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for k in row.keys():
            if k not in seen:
                seen.add(k)
                cols.append(k)

    # header
    for j, c in enumerate(cols, start=1):
        cell = ws.cell(row=start_row, column=j, value=c)
        cell.font = bold
        cell.fill = fill

    # data
    for ri, row in enumerate(rows, start=start_row + 1):
        for j, c in enumerate(cols, start=1):
            v = row.get(c)
            if isinstance(v, (dict, list)):
                v = json.dumps(v, ensure_ascii=False)
            ws.cell(row=ri, column=j, value=v)

    # autosize-ish: cap at 50 chars
    for j, c in enumerate(cols, start=1):
        sample = [str(row.get(c, "")) for row in rows[:50]]
        width = min(50, max(12, len(c), max((len(s) for s in sample), default=0) + 2))
        ws.column_dimensions[get_column_letter(j)].width = width
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)


def write_index_sheet(wb: Workbook, index_rows: list[dict]):
    ws = wb.create_sheet("Index", 0)
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="305496")
    white_bold = Font(bold=True, color="FFFFFF")

    cols = ["#", "endpoint", "group", "http_path", "params", "status", "rows", "note"]
    for j, c in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=j, value=c)
        cell.font = white_bold
        cell.fill = fill
    for ri, row in enumerate(index_rows, start=2):
        for j, c in enumerate(cols, start=1):
            ws.cell(row=ri, column=j, value=row.get(c, ""))
        if row.get("status") == "FAILURE":
            ws.cell(row=ri, column=6).font = Font(bold=True, color="C00000")
        elif row.get("status") == "SKIPPED":
            ws.cell(row=ri, column=6).font = Font(italic=True, color="808080")

    widths = [4, 28, 12, 50, 50, 10, 8, 40]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A2"


SKIPPED_DESTRUCTIVE = [
    ("changepassword",            f"{BASE}/changepassword",            "Admin",    "would change the API user's password"),
    ("immobilize",                f"{BASE}/immobilize",                "Control",  "would actually immobilize the vehicle"),
    ("mobilize",                  f"{BASE}/mobilize",                  "Control",  "would actually mobilize the vehicle"),
    ("lock",                      f"{BASE}/lock",                      "Control",  "would actually lock vehicles"),
    ("unlock",                    f"{BASE}/unlock",                    "Control",  "would actually unlock vehicles"),
    ("mosfetcontrol/enable",      f"{MOSFET_BASE}/enable",             "Mosfet",   "would actually enable mosfet"),
    ("mosfetcontrol/disable",     f"{MOSFET_BASE}/disable",            "Mosfet",   "would actually disable mosfet"),
]


def main():
    user, pw = load_env()
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    # remove the default empty sheet so insertion order is clean
    wb.remove(wb.active)

    index_rows: list[dict] = []
    now = datetime.now(timezone.utc)
    start_24h = now - timedelta(days=1)
    starttime, endtime = ms(start_24h), ms(now)
    window = f"{start_24h.isoformat(timespec='seconds')}  ->  {now.isoformat(timespec='seconds')}"

    # -------- 01 gettoken
    print(f"[01] gettoken ...")
    tok_payload = {"username": user, "password": pw}
    tok_resp = post(f"{BASE}/gettoken", tok_payload)
    rows, note = flatten_to_rows(tok_resp)
    write_sheet(wb, "01_gettoken", rows, [
        ("endpoint", "gettoken"),
        ("http_path", f"{BASE}/gettoken"),
        ("params", "username, password"),
        ("status", str(tok_resp.get("status"))),
    ])
    index_rows.append({
        "#": 1, "endpoint": "gettoken", "group": "Auth",
        "http_path": f"{BASE}/gettoken", "params": "username,password",
        "status": tok_resp.get("status") or "FAILURE",
        "rows": len(rows), "note": note or "",
    })
    if tok_resp.get("status") != "SUCCESS":
        # write what we have and abort
        write_index_sheet(wb, index_rows)
        wb.save(OUT_FILE)
        sys.exit(f"Auth failed: {tok_resp}. Partial workbook saved to {OUT_FILE}")
    token = tok_resp["data"]["token"]
    print(f"     token acquired (len={len(token)})")
    time.sleep(RATE_DELAY_S)

    # -------- 02 listvehicles (also picks sample vehicle)
    print(f"[02] listvehicles ...")
    lv = post(f"{BASE}/listvehicles", {"token": token})
    rows, note = flatten_to_rows(lv)
    write_sheet(wb, "02_listvehicles", rows, [
        ("endpoint", "listvehicles"),
        ("http_path", f"{BASE}/listvehicles"),
        ("params", "token"),
        ("status", str(lv.get("status"))),
    ])
    vehicles = [v.get("vehicleno") for v in (lv.get("data") or []) if isinstance(v, dict) and v.get("vehicleno")]
    sample_vehicle = vehicles[0] if vehicles else None
    index_rows.append({
        "#": 2, "endpoint": "listvehicles", "group": "Admin",
        "http_path": f"{BASE}/listvehicles", "params": "token",
        "status": lv.get("status") or "FAILURE", "rows": len(rows),
        "note": (note or "") + (f"  fleet={len(vehicles)} sample={sample_vehicle}" if vehicles else "  no vehicles"),
    })
    print(f"     fleet={len(vehicles)}, sample={sample_vehicle}")
    time.sleep(RATE_DELAY_S)

    # -------- catalog of remaining endpoints
    # each item: (idx, endpoint, sheet, group, url, params_dict_builder, params_summary)
    # params_dict_builder is a fn that builds the JSON body given (token, sample_vehicle)
    catalog = [
        ("listusers",               "03_listusers",               "Admin",
            f"{BASE}/listusers",
            lambda t, v: {"token": t},
            "token"),
        ("getvehicleinfo",          "04_getvehicleinfo",          "Admin",
            f"{BASE}/getvehicleinfo",
            lambda t, v: {"token": t, "vehicleno": v},
            "token,vehicleno"),
        ("getdeviceforvehicle",     "05_getdeviceforvehicle",     "Admin",
            f"{BASE}/getdeviceforvehicle",
            lambda t, v: {"token": t, "vehicleno": v},
            "token,vehicleno"),
        ("listvehicledevicemapping","06_listvehicledevicemapping","Mapping",
            f"{BASE}/listvehicledevicemapping",
            lambda t, v: {"token": t},
            "token"),
        ("getarbidparammap",        "07_getarbidparammap",        "Mapping",
            f"{BASE}/getarbidparammap",
            lambda t, v: {"token": t, "vehicleno": v},
            "token,vehicleno"),
        ("getlastgpsstatus",        "08_getlastgpsstatus",        "GPS",
            f"{BASE}/getlastgpsstatus",
            lambda t, v: {"token": t, "vehicleno": v},
            "token,vehicleno"),
        ("getgpshistory",           "09_getgpshistory",           "GPS",
            f"{BASE}/getgpshistory",
            lambda t, v: {"token": t, "vehicleno": v, "starttime": starttime, "endtime": endtime},
            f"token,vehicleno,window={window}"),
        ("getlatestcan",            "10_getlatestcan",            "CAN",
            f"{BASE}/getlatestcan",
            lambda t, v: {"token": t, "vehicleno": v},
            "token,vehicleno"),
        ("getbatterymetricshistory","11_getbatterymetricshistory","CAN",
            f"{BASE}/getbatterymetricshistory",
            lambda t, v: {"token": t, "vehicleno": v, "starttime": starttime, "endtime": endtime},
            f"token,vehicleno,window={window}"),
        ("getlastbatterymetrics",   "12_getlastbatterymetrics",   "Battery",
            f"{BASE}/getlastbatterymetrics",
            lambda t, v: {"token": t, "vehicleno": v},
            "token,vehicleno"),
        ("mosfetcontrol/currentstate","13_mosfet_currentstate",   "Mosfet",
            f"{MOSFET_BASE}/currentstate",
            lambda t, v: {"token": t, "vehicleno": v},
            "token,vehicleno"),
        ("mosfetcontrol/cmdstatus", "14_mosfet_cmdstatus",        "Mosfet",
            f"{MOSFET_BASE}/cmdstatus",
            lambda t, v: {"token": t, "vehicleno": v},
            "token,vehicleno (no cmdid - expect error-shape)"),
        ("getbletagshistory",       "15_getbletagshistory",       "BLE",
            f"{BASE}/getbletagshistory",
            lambda t, v: {"token": t, "vehicleno": v, "starttime": starttime, "endtime": endtime},
            f"token,vehicleno,window={window}"),
        ("getdistancetravelled",    "16_getdistancetravelled",    "Distance",
            f"{BASE}/getdistancetravelled",
            lambda t, v: {"token": t, "vehicleno": v, "starttime": starttime, "endtime": endtime},
            f"token,vehicleno,window={window}"),
        ("getdistancebulk",         "17_getdistancebulk",         "Distance",
            f"{BASE}/getdistancebulk",
            lambda t, v: {"token": t, "vehiclenos": [v], "starttime": starttime, "endtime": endtime},
            f"token,vehiclenos=[{sample_vehicle}],window={window}"),
        ("getfuelhistory",          "18_getfuelhistory",          "Fuel",
            f"{BASE}/getfuelhistory",
            lambda t, v: {"token": t, "vehicleno": v, "starttime": starttime, "endtime": endtime, "inlitres": False},
            f"token,vehicleno,inlitres=false,window={window}"),
        ("getfuelused",             "19_getfuelused",             "Fuel",
            f"{BASE}/getfuelused",
            lambda t, v: {"token": t, "vehicleno": v, "starttime": starttime, "endtime": endtime},
            f"token,vehicleno,window={window}"),
        ("getlastfuelstatus",       "20_getlastfuelstatus",       "Fuel",
            f"{BASE}/getlastfuelstatus",
            lambda t, v: {"token": t, "vehicleno": v},
            "token,vehicleno"),
    ]

    for i, (endpoint, sheet, group, url, build_payload, params_summary) in enumerate(catalog, start=3):
        if sample_vehicle is None and "vehicleno" in params_summary:
            print(f"[{i:02d}] {endpoint} ... SKIPPED (no sample vehicle)")
            index_rows.append({
                "#": i, "endpoint": endpoint, "group": group,
                "http_path": url, "params": params_summary,
                "status": "SKIPPED", "rows": 0,
                "note": "no vehicles in fleet",
            })
            write_sheet(wb, sheet, [], [
                ("endpoint", endpoint), ("http_path", url),
                ("params", params_summary), ("status", "SKIPPED"),
            ])
            continue

        payload = build_payload(token, sample_vehicle)
        resp = post(url, payload)
        rows, note = flatten_to_rows(resp)
        status = resp.get("status") if isinstance(resp, dict) else "FAILURE"
        print(f"[{i:02d}] {endpoint:30s} {status}  rows={len(rows)}")
        write_sheet(wb, sheet, rows, [
            ("endpoint", endpoint),
            ("http_path", url),
            ("params", params_summary),
            ("status", str(status)),
            ("err", str(resp.get("err") if isinstance(resp, dict) else "")),
            ("msg", str(resp.get("msg") if isinstance(resp, dict) else "")),
        ])
        index_rows.append({
            "#": i, "endpoint": endpoint, "group": group,
            "http_path": url, "params": params_summary,
            "status": status or "FAILURE", "rows": len(rows),
            "note": note or "",
        })
        time.sleep(RATE_DELAY_S)

    # destructive endpoints — listed in Index sheet only
    next_idx = len(catalog) + 3
    for n, (endpoint, url, group, why) in enumerate(SKIPPED_DESTRUCTIVE):
        index_rows.append({
            "#": next_idx + n, "endpoint": endpoint, "group": group,
            "http_path": url, "params": "(not called)",
            "status": "SKIPPED", "rows": 0, "note": why,
        })

    write_index_sheet(wb, index_rows)
    wb.save(OUT_FILE)
    print(f"\nWrote {OUT_FILE}")
    print(f"Sheets: {len(wb.sheetnames)} (1 Index + {len(wb.sheetnames) - 1} endpoint sheets)")


if __name__ == "__main__":
    main()
